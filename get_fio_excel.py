from pathlib import Path


def read_data_from_excel(group_list=None):
    import openpyxl

    if not group_list:
        print("ERROR: Not group list!")
        return []
    else:
        lists_path = Path('user_lists')
        wb = openpyxl.load_workbook(lists_path / group_list)
        sheet = wb.active
        rows = sheet.max_row
        list_data = [str(sheet.cell(row=1, column=1).value)]
        for row_num in range(2, rows + 1):
            list_data.append(str(sheet.cell(row=row_num, column=1).value) + ' ' + str(
                sheet.cell(row=row_num, column=2).value) + ' ' + str(sheet.cell(row=row_num, column=3).value))
        return list_data
